# genai_db_tool.py
# Streamlit GenAI Database Testing Tool (upgraded from positive.txt)
# Save as genai_db_tool.py and run: streamlit run genai_db_tool.py

import os
import re
import json
import uuid

import matplotlib.pyplot as plt
#import seaborn as sns
import time
import subprocess
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime
from dataclasses import dataclass
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import streamlit as st
import pandas as pd
import numpy as np

from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import Engine

# Optional DB drivers - import if available; show helpful errors if not.
try:
    import pymysql  # MySQL
except Exception:
    pymysql = None
try:
    import psycopg2  # Postgres
except Exception:
    psycopg2 = None
try:
    import pyodbc  # SQL Server / ODBC
except Exception:
    pyodbc = None
try:
    import oracledb  # Oracle
except Exception:
    oracledb = None

# Mongo / Hive / Databricks connectors are optional; show placeholders if missing
try:
    import pymongo
except Exception:
    pymongo = None
try:
    # pyhive for Hive if installed
    from pyhive import hive  # type: ignore
except Exception:
    hive = None

# Gemini (google generative ai) optional
try:
    import google.generativeai as genai
    from google.generativeai import GenerativeModel
except Exception:
    genai = None
    GenerativeModel = None

st.set_page_config(page_title="GenAI DB Testing Tool (Upgraded)", layout="wide")

# ---------------------------
# Helper utilities (from positive.txt)
# ---------------------------
def parse_csv_list(s: str) -> List[str]:
    return [x.strip() for x in (s or "").split(",") if x.strip()]

def read_any_file(uploaded_file) -> pd.DataFrame:
    name = (uploaded_file.name or "").lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(uploaded_file)
        elif name.endswith(".json"):
            # try lines first
            try:
                return pd.read_json(uploaded_file, lines=True)
            except Exception:
                uploaded_file.seek(0)
                return pd.read_json(uploaded_file)
        elif name.endswith(".parquet"):
            return pd.read_parquet(uploaded_file)
        elif name.endswith((".xlsx", ".xls")):
            return pd.read_excel(uploaded_file)
        else:
            # fixed width or txt
            return pd.read_fwf(uploaded_file)
    except Exception as e:
        raise RuntimeError(f"Failed to read file {uploaded_file.name}: {e}")

def normalize_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    for c in df2.columns:
        if pd.api.types.is_datetime64_any_dtype(df2[c]):
            df2[c] = pd.to_datetime(df2[c], errors='coerce')
        elif pd.api.types.is_numeric_dtype(df2[c]):
            pass
        else:
            df2[c] = df2[c].astype(str).str.strip()
    return df2

def align_and_diff(left: pd.DataFrame, right: pd.DataFrame, join_keys: List[str],
                   compare_cols: Optional[List[str]] = None, numeric_tolerance: float = 0.0):
    l = normalize_dtypes(left)
    r = normalize_dtypes(right)
    if not join_keys:
        common = sorted(list(set(l.columns).intersection(set(r.columns))))
        if not common:
            raise ValueError("No join keys provided and no common columns to join on.")
        join_keys = common
    for k in join_keys:
        if k not in l.columns or k not in r.columns:
            raise ValueError(f"Join key '{k}' not present in both datasets.")
    l_key = l.set_index(join_keys, drop=False)
    r_key = r.set_index(join_keys, drop=False)
    only_left = l_key.loc[~l_key.index.isin(r_key.index)].reset_index(drop=True)
    only_right = r_key.loc[~r_key.index.isin(l_key.index)].reset_index(drop=True)
    common_left = l_key.loc[l_key.index.isin(r_key.index)]
    common_right = r_key.loc[r_key.index.isin(l_key.index)].loc[common_left.index]

    if compare_cols:
        cols = [c for c in compare_cols if c in common_left.columns and c in common_right.columns]
    else:
        cols = sorted(list(set(common_left.columns).intersection(set(common_right.columns))))
        cols = [c for c in cols if c not in join_keys]

    records = []
    for c in cols:
        ls = common_left[c]
        rs = common_right[c]
        if pd.api.types.is_numeric_dtype(ls) and pd.api.types.is_numeric_dtype(rs) and numeric_tolerance > 0:
            eq = (ls - rs).abs() <= numeric_tolerance
        else:
            eq = ls.fillna("__NULL__").astype(str) == rs.fillna("__NULL__").astype(str)
        mask = ~eq
        if mask.any():
            df_diff = pd.DataFrame({**{k: common_left[k] for k in join_keys},
                                    "column": c,
                                    "left_value": ls.astype(object),
                                    "right_value": rs.astype(object)})
            df_diff = df_diff[mask]
            records.append(df_diff)
    value_diffs = pd.concat(records, ignore_index=True) if records else pd.DataFrame(
        columns=(join_keys + ["column", "left_value", "right_value"]))
    return only_left, only_right, value_diffs

def export_mapping_to_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"

    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data and apply color
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Color the row based on match column
            if "match" in df.columns:
                match_val = getattr(row, "match")
                if match_val is True:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
                else:
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # amber
                cell.fill = fill

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# ---------------------------
# DB connection helpers (extended)
# ---------------------------
def build_conn_url(
        db_type: str,
        host: str,
        port: str,
        user: str,
        password: str,
        database: str,
        sid_service: str = "",
        aws_access_key_id: str = "",
        aws_secret_access_key: str = "",
        aws_session_token: str = "",
        region: str = "",
        s3_staging_dir: str = "",
        http_path: str = ""  # New param for Databricks
) -> str:
    if db_type == "MySQL":
        return f"mysql+pymysql://{user}:{password}@{host}:{port}/{database}"
    if db_type == "Postgres":
        return f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}"
    if db_type == "SQL Server":
        driver = "ODBC+Driver+17+for+SQL+Server"
        return f"mssql+pyodbc://{user}:{password}@{host}:{port}/{database}?driver={driver}"
    if db_type == "Oracle":
        if sid_service:
            return f"oracle+oracledb://{user}:{password}@{host}:{port}/?service_name={sid_service}"
        else:
            return f"oracle+oracledb://{user}:{password}@{host}:{port}/{database}"
    if db_type == "Redshift":
        return f"redshift+psycopg2://{user}:{password}@{host}:{port}/{database}"
    if db_type == "Athena":
        if not (aws_access_key_id and aws_secret_access_key and s3_staging_dir and region):
            raise ValueError("Athena requires AWS Access Key ID, Secret Access Key, Region, and S3 staging directory.")
        token_part = f"&aws_session_token={aws_session_token}" if aws_session_token else ""
        s3_dir_enc = s3_staging_dir if s3_staging_dir.endswith("/") else s3_staging_dir + "/"
        return (
            f"awsathena+rest://{aws_access_key_id}:{aws_secret_access_key}@{host}:{port}/{database}"
            f"?s3_staging_dir={s3_dir_enc}&region_name={region}{token_part}"
        )
    if db_type.lower() == "databricks":
        if not (host and http_path and password):
            raise ValueError("Databricks requires Workspace Hostname, HTTP Path, and Personal Access Token.")
       # http_path_enc = http_path if http_path.startswith("/") else "/" + http_path
        return (
            f"databricks://token:{password}@{host}?http_path={http_path}"
        )
    raise ValueError("Unsupported DB type")

def try_connect(url: str) -> Tuple[Optional[Engine], Optional[str]]:
    try:
        engine = create_engine(url, pool_pre_ping=True)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return engine, None
    except Exception as e:
        if "athena" in url:
            return engine, None
        return None, str(e)

# ---------------------------
# Gemini / generative -> SQL helper (renamed & improved)
# Returns only SQL string (no markdown/fences)
# ---------------------------
def configure_gemini_from_key(key: str) -> Tuple[bool, Optional[str]]:
    if genai is None:
        return False, "google-generativeai package not installed"
    try:
        genai.configure(api_key=key)
        return True, None
    except Exception as e:
        return False, str(e)

def _clean_generated_sql_text(text: str) -> str:
    if not text:
        return ""
    t = text.strip()
    t = re.sub(r"^```(?:sql|json)?\s*", "", t, flags=re.IGNORECASE | re.MULTILINE)
    t = re.sub(r"```$", "", t, flags=re.MULTILINE)
    try:
        m = re.search(r"(\{[\s\S]*\})", t)
        if m:
            try:
                d = json.loads(m.group(1))
                for k in ("source_sql", "target_sql", "query", "sql"):
                    if k in d and isinstance(d[k], str) and d[k].strip():
                        return d[k].strip()
            except Exception:
                pass
    except Exception:
        pass
    t = re.sub(r"^[\s`]*sql[:\s]+", "", t, flags=re.IGNORECASE)
    t = t.strip().strip('`').strip()
    t = re.sub(r"^```|```$", "", t).strip()
    lines = t.splitlines()
    sql_start_idx = 0
    for i, ln in enumerate(lines):
        if re.search(r"\b(select|with|insert|update|delete|create|drop)\b", ln, flags=re.I):
            sql_start_idx = i
            break
    lines = lines[sql_start_idx:]
    t = "\n".join(lines).strip()
    t = re.sub(r"^['\"]{3}", "", t)
    t = re.sub(r"['\"]{3}$", "", t)
    return t.strip()

def generate_sql(prompt: str, schema_text: str = "", model_name: str = "gemini-2.5-flash") -> str:
    if genai is None or GenerativeModel is None:
        raise RuntimeError("Gemini (google.generativeai) not installed; cannot generate SQL. Please install google-generativeai or provide manual SQL.")
    system = (
        "You are a SQL generation assistant. Use the exact schema the user provided. "
        "Return only the SQL query (no markdown, no explanation). If you return JSON, ensure the 'sql' or 'source_sql' key contains the SQL string."
    )
    user_message = f"User prompt:\n{prompt}\n\nSchema:\n{schema_text}\n\nReturn only SQL, no commentary."
    model = GenerativeModel(model_name)
    try:
        resp = model.generate_content([{"role": "user", "parts": [system]}, {"role": "user", "parts": [user_message]}])
        raw = getattr(resp, "text", "") or str(resp)
        sql = _clean_generated_sql_text(raw)
        return sql
    except Exception as e:
        try:
            resp2 = model.generate_content(user_message)
            raw2 = getattr(resp2, "text", "") or str(resp2)
            sql2 = _clean_generated_sql_text(raw2)
            return sql2
        except Exception as e2:
            raise RuntimeError(f"Gemini generation failed: {e} / fallback failed: {e2}")

def _log_entry(section: str, prompt: str, generated_sql: str = "", executed_sql: str = "",
               result_summary: Dict | None = None):
    logs = st.session_state.get("logs", [])
    logs.append({
        "ts": datetime.utcnow().isoformat(),
        "section": section,
        "prompt": prompt,
        "generated_sql": generated_sql,
        "executed_sql": executed_sql,
        "result_summary": result_summary or {}
    })
    st.session_state["logs"] = logs

def log(section: str, info: dict):
    _log_entry(section, prompt="", result_summary=info)

def run_dbt_command(project_dir: str, args: List[str]) -> Tuple[int, str, str]:
    cmd = ["dbt", "--project-dir", project_dir] + args
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
        return proc.returncode, proc.stdout, proc.stderr
    except Exception as e:
        raise RuntimeError(f"Failed to run dbt command: {e}")

@dataclass
class DBConfig:
    kind: str
    host: str
    port: str
    database: str
    username: str
    password: str

def execute_sql(cfg: DBConfig, sql: str) -> pd.DataFrame:
    url = build_conn_url(
        db_type=cfg.kind,
        host=cfg.host,
        port=cfg.port,
        user=cfg.username,
        password=cfg.password,
        database=cfg.database
    )
    eng, err = try_connect(url)
    if not eng:
        raise RuntimeError(f"DB connection failed: {err}")
    return pd.read_sql(text(sql), eng)

# ---------------------------
# Left explorer navigation (keeps sections in left)
# ---------------------------
st.sidebar.title("Explorer")
section = st.sidebar.radio("Select section", [
    "Single Database", "Multiple Database Comparison", "FlatFile ↔ DB", "SCD Validations",
    "Mapping Generator", "Data Quality", "Reconciliation", "Profiling", "DBT", "Logs & Downloads"
], index=0)

# Top-level Gemini API key in sidebar (stable)
st.sidebar.markdown("---")
gemini_key = st.sidebar.text_input("Gemini API Key (optional)", type="password", key="gemini_key_input")
if gemini_key:
    ok, err = configure_gemini_from_key(gemini_key)
    if ok:
        st.sidebar.success("Gemini configured")
        st.session_state["gemini_key"] = gemini_key
    else:
        st.sidebar.error(f"Gemini error: {err}")

if section == "Single Database":
    st.header("Single Database — NL→SQL + Run")
    st.markdown("Connect to a single database, ask using natural language (Gemini), generate SQL and run it. Logs saved automatically.")

    col1, col2, col3 = st.columns(3)
    db_type = col1.selectbox("DB Type", ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"], key="single_db_type")

    if db_type == "Athena":
        st.markdown("### Athena Connection Details")
        aws_access_key_id = st.text_input("AWS Access Key ID", type="password", key="athena_key")
        aws_secret_access_key = st.text_input("AWS Secret Access Key", type="password", key="athena_secret")
        aws_session_token = st.text_input("AWS Session Token (optional)", type="password", key="athena_token")
        region = st.text_input("AWS Region", value="us-east-1", key="athena_region")
        s3_staging_dir = st.text_input("S3 Staging Dir (s3://bucket/folder/)", key="athena_s3")
        database = st.text_input("Database / Catalog", key="single_db")
        host = f"athena.{region}.amazonaws.com"
        port = "443"
        user = aws_access_key_id
        pwd = aws_secret_access_key
        sid_service = s3_staging_dir
        http_path = ""
    elif db_type == "Databricks":
        st.markdown("### Databricks Connection Details")
        databricks_host = st.text_input("Workspace Hostname", key="single_db_databricks_host")
        databricks_http_path = st.text_input("HTTP Path", key="single_db_databricks_http_path")
        databricks_token = st.text_input("Personal Access Token", type="password", key="single_db_databricks_token")
        database = st.text_input("Database / Catalog", key="single_db")
        host = databricks_host
        port = "443"
        user = "token"
        pwd = databricks_token
        sid_service = ""
        aws_access_key_id = ""
        aws_secret_access_key = ""
        aws_session_token = ""
        region = ""
        s3_staging_dir = ""
        http_path = databricks_http_path
    else:
        host = col1.text_input("Host", key="single_host")
        port = col2.text_input("Port", value=("3306" if db_type == "MySQL" else "5432" if db_type == "Postgres" else "1433" if db_type == "SQL Server" else "1521"), key="single_port")
        user = col1.text_input("User", key="single_user")
        pwd = col2.text_input("Password", type="password", key="single_pwd")
        database = col3.text_input("Database / Catalog", key="single_db")
        sid_service = col3.text_input("SID/Service (Oracle)", key="single_sid")
        aws_access_key_id = ""
        aws_secret_access_key = ""
        aws_session_token = ""
        region = ""
        s3_staging_dir = ""
        http_path = ""

    if st.button("Connect to DB", key="single_connect"):
        try:
            url = build_conn_url(
                db_type, host, port, user, pwd, database, sid_service,
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                aws_session_token=aws_session_token,
                region=region,
                s3_staging_dir=s3_staging_dir,
                http_path=http_path
            )
            eng, err = try_connect(url)
            if eng:
                st.success("Connected to DB")
                st.session_state["single_engine"] = eng
                st.session_state["single_conn_url"] = url
            else:
                st.error(f"Connection failed: {err}")
        except Exception as e:
            st.error(f"Connection error: {e}")

    engine = st.session_state.get("single_engine")
    schema_text = ""
    if engine:
        try:
            inspector = inspect(engine)
            parts = []
            for t in inspector.get_table_names()[:50]:
                cols = inspector.get_columns(t)
                parts.append(f"{t}: " + ", ".join([f"{c['name']}({c['type']})" for c in cols]))
            schema_text = "\n".join(parts)
            st.text_area("Detected schema (read-only)", value=schema_text, height=200)
        except Exception:
            schema_text = ""

    nl = st.text_area("Ask in plain English (generate SQL)", key="single_nl", height=140)
    gen_model = st.selectbox("Model (Gemini)", ["gemini-2.5-flash"], key="single_gen_model")
    colA, colB = st.columns([1, 1])
    if colA.button("Generate SQL (Gemini)", key="single_generate"):
        if not nl.strip():
            st.warning("Enter a prompt")
        else:
            if "gemini_key" not in st.session_state:
                st.error("Provide Gemini API key in sidebar to use auto-generation.")
            else:
                try:
                    sql = generate_sql(nl, schema_text, model_name=gen_model)
                    st.session_state["single_generated_sql"] = sql
                    st.code(sql, language="sql")
                    _log_entry("Single Database", nl, generated_sql=sql)
                    st.success("SQL generated and logged.")
                except Exception as e:
                    st.error(f"SQL generation failed: {e}")

    # Manual override textbox
    sql_manual = st.text_area("Manual/Generated SQL (edit if needed)",
                              value=st.session_state.get("single_generated_sql", ""), height=180, key="single_sql_text")
    if colB.button("Run SQL", key="single_run"):
        if not engine:
            st.error("Connect to DB first")
        else:
            if not sql_manual.strip():
                st.error("Provide SQL to run")
            else:
                try:
                    df = pd.read_sql(text(sql_manual), engine)
                    st.session_state["single_last_result"] = df
                    st.success(f"Query returned {len(df)} rows and {len(df.columns)} columns.")
                    st.dataframe(df.head(500), use_container_width=True)
                    # summarize
                    summary = {"rows": len(df), "cols": len(df.columns), "columns": list(df.columns[:50])}
                    _log_entry("Single Database", nl or "",
                               generated_sql=st.session_state.get("single_generated_sql", ""), executed_sql=sql_manual,
                               result_summary=summary)
                    # pie chart example: if user asked to group by some column, we allow them to pick a column
                    if len(df.columns) > 0:
                        col_for_pie = st.selectbox("Optional: pick a column for pie chart (freq)",
                                                   options=list(df.columns), index=0)
                        vc = df[col_for_pie].astype(str).value_counts().head(10)
                        st.subheader("Pie chart — top values")
                        st.pyplot(vc.plot.pie(autopct="%1.1f%%").get_figure())
                except Exception as e:
                    st.error(f"SQL execution error: {e}")

# ---------------------------
# Section: Multiple Database Comparison
# ---------------------------
elif section == "Multiple Database Comparison":
    st.header("Multiple Database Comparison (Source ↔ Target)")
    st.markdown("Connect to Source and Target DBs and run schema/data/count/duplicate checks.")

    # Source connection block
    st.subheader("Source Connection")
    s_c1, s_c2, s_c3 = st.columns(3)
    s_db_type = s_c1.selectbox("Source DB Type",
                               ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
                               key="cmp_src_type")
    if s_db_type == "Athena":
        st.markdown("#### Athena Source Details")
        s_aws_access_key_id = st.text_input("AWS Access Key ID (Source)", type="password", key="cmp_src_athena_key")
        s_aws_secret_access_key = st.text_input("AWS Secret Access Key (Source)", type="password",
                                                key="cmp_src_athena_secret")
        s_aws_session_token = st.text_input("AWS Session Token (Source, optional)", type="password",
                                            key="cmp_src_athena_token")
        s_region = st.text_input("AWS Region (Source)", value="us-east-1", key="cmp_src_athena_region")
        s3_staging_dir = st.text_input("S3 Staging Dir (Source)", key="cmp_src_athena_s3")
        s_host = f"athena.{s_region}.amazonaws.com"
        s_port = "443"
        s_user = s_aws_access_key_id
        s_pwd = s_aws_secret_access_key
        s_db = st.text_input("Database (Source)", key="cmp_src_db")
        s_sid = s3_staging_dir
    elif s_db_type == "Databricks":
        st.markdown("### Databricks Connection Details")
        databricks_host = st.text_input("Workspace Hostname", key="single_db_databricks_host")
        databricks_http_path = st.text_input("HTTP Path", key="single_db_databricks_http_path")
        databricks_token = st.text_input("Personal Access Token", type="password", key="single_db_databricks_token")
        database = st.text_input("Database / Catalog", key="single_db")
        s_host = databricks_host
        s_port = "443"
        s_user = "token"
        s_pwd = databricks_token
        s_sid_service = ""
        s_sid = ""
        s_aws_access_key_id = ""
        s_aws_secret_access_key = ""
        s_aws_session_token = ""
        s_region = ""
        s3_staging_dir = ""
        http_path = databricks_http_path
        s_db = st.text_input("Database (Source)", key="cmp_src_db")
    else:
        s_host = s_c1.text_input("Host", key="cmp_src_host")
        s_port = s_c2.text_input("Port", value=("3306" if s_db_type == "MySQL" else "5432"), key="cmp_src_port")
        s_user = s_c1.text_input("User", key="cmp_src_user")
        s_pwd = s_c2.text_input("Password", type="password", key="cmp_src_pwd")
        s_db = s_c3.text_input("Database", key="cmp_src_db")
        s_sid = s_c3.text_input("SID/Service (Oracle)", key="cmp_src_sid")
        s_aws_access_key_id = ""
        s_aws_secret_access_key = ""
        s_aws_session_token = ""
        s_region = ""
        s3_staging_dir = ""

    if s_c3.button("Connect Source", key="cmp_src_connect"):
        try:
            url = build_conn_url(
                s_db_type, s_host, s_port, s_user, s_pwd, s_db, s_sid,
                aws_access_key_id=s_aws_access_key_id,
                aws_secret_access_key=s_aws_secret_access_key,
                aws_session_token=s_aws_session_token,
                region=s_region,
                s3_staging_dir=s3_staging_dir,
                http_path = databricks_http_path
            )
            eng, err = try_connect(url)
            if eng:
                st.session_state["cmp_src_engine"] = eng
                st.success("Source connected")
            else:
                st.error(f"Source connect failed: {err}")
        except Exception as e:
            st.error(f"Source connect error: {e}")

    # Target connection block
    st.subheader("Target Connection")
    t_c1, t_c2, t_c3 = st.columns(3)
    t_db_type = t_c1.selectbox("Target DB Type",
                               ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
                               key="cmp_tgt_type")
    if t_db_type == "Athena":
        st.markdown("#### Athena Target Details")
        t_aws_access_key_id = st.text_input("AWS Access Key ID (Target)", type="password", key="cmp_tgt_athena_key")
        t_aws_secret_access_key = st.text_input("AWS Secret Access Key (Target)", type="password",
                                                key="cmp_tgt_athena_secret")
        t_aws_session_token = st.text_input("AWS Session Token (Target, optional)", type="password",
                                            key="cmp_tgt_athena_token")
        t_region = st.text_input("AWS Region (Target)", value="us-east-1", key="cmp_tgt_athena_region")
        t3_staging_dir = st.text_input("S3 Staging Dir (Target)", key="cmp_tgt_athena_s3")
        t_host = f"athena.{t_region}.amazonaws.com"
        t_port = "443"
        t_user = t_aws_access_key_id
        t_pwd = t_aws_secret_access_key
        t_db = st.text_input("Database (Target)", key="cmp_tgt_db")
        t_sid = t3_staging_dir
    elif t_db_type == "Databricks":
        st.markdown("### Databricks Connection Details")
        databricks_host = st.text_input("Workspace Hostname", key="t_single_db_databricks_host")
        databricks_http_path = st.text_input("HTTP Path", key="t_single_db_databricks_http_path")
        databricks_token = st.text_input("Personal Access Token", type="password", key="t_single_db_databricks_token")
        database = st.text_input("Database / Catalog", key="t_single_db")
        t_host = databricks_host
        t_port = "443"
        t_user = "token"
        t_pwd = databricks_token
        t_sid_service = ""
        t_sid = ""
        t_aws_access_key_id = ""
        t_aws_secret_access_key = ""
        t_aws_session_token = ""
        t_region = ""
        t3_staging_dir = ""
        http_path = databricks_http_path
        t_db = st.text_input("Database (Source)", key="cmp_tgt_db")
    else:
        t_host = t_c1.text_input("Host", key="cmp_tgt_host")
        t_port = t_c2.text_input("Port", value=("3306" if t_db_type == "MySQL" else "5432"), key="cmp_tgt_port")
        t_user = t_c1.text_input("User", key="cmp_tgt_user")
        t_pwd = t_c2.text_input("Password", type="password", key="cmp_tgt_pwd")
        t_db = t_c3.text_input("Database", key="cmp_tgt_db")
        t_sid = t_c3.text_input("SID/Service (Oracle)", key="cmp_tgt_sid")
        t_aws_access_key_id = ""
        t_aws_secret_access_key = ""
        t_aws_session_token = ""
        t_region = ""
        t3_staging_dir = ""

    if t_c3.button("Connect Target", key="cmp_tgt_connect"):
        try:
            url = build_conn_url(
                t_db_type, t_host, t_port, t_user, t_pwd, t_db, t_sid,
                aws_access_key_id=t_aws_access_key_id,
                aws_secret_access_key=t_aws_secret_access_key,
                aws_session_token=t_aws_session_token,
                region=t_region,
                s3_staging_dir=t3_staging_dir,
                http_path=databricks_http_path
            )
            eng, err = try_connect(url)
            if eng:
                st.session_state["cmp_tgt_engine"] = eng
                st.success("Target connected")
            else:
                st.error(f"Target connect failed: {err}")
        except Exception as e:
            st.error(f"Target connect error: {e}")

    src_engine = st.session_state.get("cmp_src_engine")
    tgt_engine = st.session_state.get("cmp_tgt_engine")

    st.markdown("---")
    check_type = st.selectbox("Type of check",
                              ["Schema comparison", "Data comparison (sample)", "Count comparison", "Duplicate checks",
                               "Aggregated grouping"], index=0)
    src_table = st.text_input("Source table name (or comma-separated list)", key="cmp_src_table")
    tgt_table = st.text_input("Target table name (or comma-separated list)", key="cmp_tgt_table")
    join_keys = parse_csv_list(st.text_input("Join keys (comma-separated)", key="cmp_joinkeys"))
    compare_columns = parse_csv_list(st.text_input("Columns to compare (comma-separated)", key="cmp_comparecols"))

    if st.button("Run Check", key="cmp_run"):
        if check_type == "Schema comparison":
            if not src_engine or not tgt_engine:
                st.error("Connect both Source and Target first.")
            else:
                try:
                    # Helper function to get columns for Databricks using DESCRIBE TABLE
                    def get_databricks_columns(engine, full_table_name):
                        sql = f"DESCRIBE TABLE {full_table_name}"
                        df = pd.read_sql(text(sql), engine)
                        # Return dict {col_name: data_type}
                        return {row["col_name"]: row["data_type"] for _, row in df.iterrows()}


                    # Helper function to get columns for other DBs using inspector
                    def get_otherdb_columns(engine, table_name):
                        insp = inspect(engine)
                        cols = insp.get_columns(table_name)
                        return {c["name"]: str(c["type"]) for c in cols}


                    src_tables = [t.strip() for t in src_table.split(",") if t.strip()]
                    tgt_tables = [t.strip() for t in tgt_table.split(",") if t.strip()]
                    schema_report = []

                    for s_table, t_table in zip(src_tables, tgt_tables):
                        # Determine source columns
                        if s_db_type.lower() == "databricks":
                            s_cols = get_databricks_columns(src_engine, s_table)
                        else:
                            s_cols = get_otherdb_columns(src_engine, s_table)

                        # Determine target columns
                        if t_db_type.lower() == "databricks":
                            t_cols = get_databricks_columns(tgt_engine, t_table)
                        else:
                            t_cols = get_otherdb_columns(tgt_engine, t_table)

                        diff_cols = []
                        for k in sorted(set(list(s_cols.keys()) + list(t_cols.keys()))):
                            diff_cols.append({
                                "col": k,
                                "src_type": s_cols.get(k, ""),
                                "tgt_type": t_cols.get(k, ""),
                                "match": s_cols.get(k, "") == t_cols.get(k, "")
                            })
                        schema_report.append({"src_table": s_table, "tgt_table": t_table, "columns": diff_cols})

                    st.success("Schema comparison complete")
                    # display a simple summary table for each pair
                    for r in schema_report:
                        st.subheader(f"{r['src_table']} ↔ {r['tgt_table']}")
                        df_rows = pd.DataFrame(r["columns"])
                        st.dataframe(df_rows, use_container_width=True)
                        st.download_button(f"Download schema diff {r['src_table']}",
                                           df_rows.to_csv(index=False).encode("utf-8"),
                                           file_name=f"schema_diff_{r['src_table']}.csv")
                    _log_entry("Multiple DB Comparison", f"Schema compare {src_table} vs {t_table}",
                               result_summary={"pairs": len(schema_report)})
                except Exception as e:
                    st.error(f"Schema comparison failed: {e}")
        elif check_type == "Count comparison":
            if not src_engine or not tgt_engine:
                st.error("Connect both Source and Target first.")
            else:
                try:
                    s_q = f"SELECT COUNT(*) as cnt FROM {src_table}"
                    t_q = f"SELECT COUNT(*) as cnt FROM {tgt_table}"
                    s_cnt = pd.read_sql(text(s_q), src_engine).iloc[0]["cnt"]
                    t_cnt = pd.read_sql(text(t_q), tgt_engine).iloc[0]["cnt"]
                    st.metric("Source count", int(s_cnt))
                    st.metric("Target count", int(t_cnt))
                    st.info("MISMATCH" if int(s_cnt) != int(t_cnt) else "MATCH")
                    _log_entry("Multiple DB Comparison", f"Count compare {src_table} vs {tgt_table}",
                               result_summary={"src_count": int(s_cnt), "tgt_count": int(t_cnt)})
                except Exception as e:
                    st.error(f"Count comparison failed: {e}")

        elif check_type == "Data comparison (sample)":
            if not src_engine or not tgt_engine:
                st.error("Connect both Source and Target first.")
            else:
                try:
                    s_sql = f"SELECT * FROM {src_table} LIMIT 1000"
                    t_sql = f"SELECT * FROM {tgt_table} LIMIT 1000"
                    left = pd.read_sql(text(s_sql), src_engine)
                    right = pd.read_sql(text(t_sql), tgt_engine)
                    only_l, only_r, diffs = align_and_diff(left, right, join_keys, compare_columns)
                    st.metric("Only in source", len(only_l))
                    st.metric("Only in target", len(only_r))
                    st.metric("Value diffs", len(diffs))
                    st.dataframe(diffs.head(500), use_container_width=True)
                    _log_entry("Multiple DB Comparison", f"Data compare {src_table} vs {tgt_table}",
                               result_summary={"only_src": len(only_l), "only_tgt": len(only_r),
                                               "value_diffs": len(diffs)})
                except Exception as e:
                    st.error(f"Data comparison failed: {e}")

        elif check_type == "Duplicate checks":
            # run duplicate detection on source and target and show
            if not src_engine or not tgt_engine:
                st.error("Connect both Source and Target first.")
            else:
                try:
                    s_df = pd.read_sql(text(f"SELECT * FROM {src_table} LIMIT 20000"), src_engine)
                    t_df = pd.read_sql(text(f"SELECT * FROM {tgt_table} LIMIT 20000"), tgt_engine)
                    s_dup = s_df.duplicated(keep=False).sum()
                    t_dup = t_df.duplicated(keep=False).sum()
                    st.metric("Source duplicate rows", int(s_dup))
                    st.metric("Target duplicate rows", int(t_dup))
                    _log_entry("Multiple DB Comparison", f"Duplicate check {src_table} vs {tgt_table}",
                               result_summary={"src_dup": int(s_dup), "tgt_dup": int(t_dup)})
                except Exception as e:
                    st.error(f"Duplicate check failed: {e}")

        elif check_type == "Aggregated grouping":
            # user enters grouping expresssion in a single text area (e.g., sum(amount) by country|state)
            agg_expr = st.text_input("Aggregate expression (e.g. sum(amount) by country|state)", key="cmp_agg_expr")
            if not agg_expr:
                st.warning("Provide aggregate expression")
            else:
                try:
                    s_q = f"SELECT {agg_expr} FROM {src_table}"
                    t_q = f"SELECT {agg_expr} FROM {tgt_table}"
                    s_agg = pd.read_sql(text(s_q), src_engine)
                    t_agg = pd.read_sql(text(t_q), tgt_engine)
                    st.subheader("Source aggregate (sample)")
                    st.dataframe(s_agg.head(200), use_container_width=True)
                    st.subheader("Target aggregate (sample)")
                    st.dataframe(t_agg.head(200), use_container_width=True)
                    _log_entry("Multiple DB Comparison", f"Aggregate {agg_expr} {src_table} vs {tgt_table}",
                               result_summary={"src_agg_rows": len(s_agg), "tgt_agg_rows": len(t_agg)})
                except Exception as e:
                    st.error(f"Aggregate execution failed: {e}")

# ---------------------------
# Section: FlatFile ↔ DB
# ---------------------------
elif section == "FlatFile ↔ DB":
    st.header("Flat File / S3 Bucket → Database Comparison")
    st.markdown("Upload file (CSV/XLSX/Parquet/JSON/TXT) and compare with a DB table.")

    # DB connect (stable inputs)
    c1, c2, c3 = st.columns(3)
    ff_db_type = c1.selectbox("DB Type",
                              ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
                              key="ff_dbtype")
    if ff_db_type == "Athena":
        st.markdown("#### Athena Connection Details")
        aws_access_key_id = st.text_input("AWS Access Key ID", type="password", key="ff_athena_key")
        aws_secret_access_key = st.text_input("AWS Secret Access Key", type="password", key="ff_athena_secret")
        aws_session_token = st.text_input("AWS Session Token (optional)", type="password", key="ff_athena_token")
        region = st.text_input("AWS Region", value="us-east-1", key="ff_athena_region")
        s3_staging_dir = st.text_input("S3 Staging Dir (s3://bucket/folder/)", key="ff_athena_s3")
        ff_host = f"athena.{region}.amazonaws.com"
        ff_port = "443"
        ff_user = aws_access_key_id
        ff_pwd = aws_secret_access_key
        ff_db = st.text_input("Database", key="ff_db")
        ff_sid = s3_staging_dir
    elif ff_db_type == "Databricks":
        st.markdown("### Databricks Connection Details")
        databricks_host = st.text_input("Workspace Hostname", key="single_db_databricks_host")
        databricks_http_path = st.text_input("HTTP Path", key="single_db_databricks_http_path")
        databricks_token = st.text_input("Personal Access Token", type="password", key="single_db_databricks_token")
        database = st.text_input("Database / Catalog", key="single_db")
        ff_host = databricks_host
        ff_port = "443"
        ff_user = "token"
        ff_pwd = databricks_token
        sid_service = ""
        aws_access_key_id = ""
        aws_secret_access_key = ""
        aws_session_token = ""
        region = ""
        s3_staging_dir = ""
        ff_db = st.text_input("Database", key="ff_db")
        ff_sid = ""
        http_path = databricks_http_path
    else:
        ff_host = c1.text_input("Host", key="ff_host")
        ff_port = c2.text_input("Port", value="3306", key="ff_port")
        ff_user = c1.text_input("User", key="ff_user")
        ff_pwd = c2.text_input("Password", type="password", key="ff_pwd")
        ff_db = c3.text_input("Database", key="ff_db")
        ff_sid = c3.text_input("SID/Service (Oracle)", key="ff_sid")
        aws_access_key_id = ""
        aws_secret_access_key = ""
        aws_session_token = ""
        region = ""
        s3_staging_dir = ""

    # Define the target table name or SQL query input
    ff_table_or_sql = st.text_input("Target Table name or SELECT SQL", key="ff_table_sql")

    if c3.button("Connect DB (file compare)", key="ff_connect"):
        try:
            url = build_conn_url(
                ff_db_type, ff_host, ff_port, ff_user, ff_pwd, ff_db, ff_sid,
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                aws_session_token=aws_session_token,
                region=region,
                s3_staging_dir=s3_staging_dir,
                http_path=databricks_http_path
            )
            eng, err = try_connect(url)
            if eng:
                st.session_state["ff_engine"] = eng
                st.success("DB connected for file compare")
            else:
                st.error(f"Connect failed: {err}")
        except Exception as e:
            st.error(f"Connect error: {e}")

    uploaded = st.file_uploader("Upload file (CSV/JSON/Parquet/XLSX/TXT)",
                                type=["csv", "json", "parquet", "xlsx", "xls", "txt"], key="ff_upload")
    join_keys = parse_csv_list(st.text_input("Join keys (comma separated)", key="ff_joinkeys"))
    compare_cols = parse_csv_list(st.text_input("Compare columns (comma separated)", key="ff_comparecols"))
    if st.button("Run File ↔ DB Comparison", key="ff_run"):
        eng = st.session_state.get("ff_engine")
        if not eng:
            st.error("Connect the DB first")
        elif uploaded is None:
            st.error("Upload a file")
        elif not ff_table_or_sql.strip():
            st.error("Provide target table name or SELECT SQL")
        else:
            try:
                file_df = read_any_file(uploaded)
                if ff_table_or_sql.strip().lower().startswith("select"):
                    db_df = pd.read_sql(text(ff_table_or_sql), eng)
                else:
                    db_df = pd.read_sql(text(f"SELECT * FROM {ff_table_or_sql}"), eng)
                only_l, only_r, diffs = align_and_diff(db_df, file_df, join_keys, compare_cols)
                st.metric("Only in DB", len(only_l))
                st.metric("Only in File", len(only_r))
                st.metric("Value diffs", len(diffs))
                st.dataframe(diffs.head(500))
                _log_entry("File↔DB", f"File {uploaded.name} vs {ff_table_or_sql}",
                           result_summary={"only_db": len(only_l), "only_file": len(only_r), "diffs": len(diffs)})
            except Exception as e:
                st.error(f"File↔DB compare failed: {e}")

# ---------------------------
# Section: SCD Validations (basic placeholders)
# ---------------------------
elif section == "SCD Validations":
    st.header("🔁 SCD Type Validations")
    st.info("Run SCD Type 1 and Type 2 validations against a connected DB table or uploaded files.")

    # Choose source type: DB or File upload for old and new snapshots
    scd_source_type = st.radio("Source type", ["Database", "File Upload"], horizontal=True, key="scd_source_type")

    # Initialize dataframes in session state if not present
    if "scd_df_old" not in st.session_state:
        st.session_state["scd_df_old"] = None
    if "scd_df_new" not in st.session_state:
        st.session_state["scd_df_new"] = None

    df_old = st.session_state["scd_df_old"]
    df_new = st.session_state["scd_df_new"]

    if scd_source_type == "Database":
        # DB connection inputs
        scd_kind = st.selectbox("DB kind",
                                ["PostgreSQL", "Databricks", "SQL Server", "MySQL", "Hive", "Oracle", "Redshift",
                                 "Athena"], key="scd_kind")
        if scd_kind == "Athena":
            st.markdown("### Athena Connection Details")
            aws_access_key_id = st.text_input("AWS Access Key ID", type="password", key="scd_athena_key")
            aws_secret_access_key = st.text_input("AWS Secret Access Key", type="password", key="scd_athena_secret")
            aws_session_token = st.text_input("AWS Session Token (optional)", type="password", key="scd_athena_token")
            region = st.text_input("AWS Region", value="us-east-1", key="scd_athena_region")
            s3_staging_dir = st.text_input("S3 Staging Dir (s3://bucket/folder/)", key="scd_athena_s3")
            host = f"athena.{region}.amazonaws.com"
            port = "443"
            user = aws_access_key_id
            pwd = aws_secret_access_key
            sid_service = s3_staging_dir
        elif scd_kind == "Databricks":
            st.markdown("### Databricks Connection Details")
            databricks_host = st.text_input("Workspace Hostname", key="single_db_databricks_host")
            databricks_http_path = st.text_input("HTTP Path", key="single_db_databricks_http_path")
            databricks_token = st.text_input("Personal Access Token", type="password", key="single_db_databricks_token")
            database = st.text_input("Database / Catalog", key="single_db")
            host = databricks_host
            port = "443"
            user = "token"
            pwd = databricks_token
            sid_service = ""
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""
            http_path = databricks_http_path
        else:
            host = st.text_input("Host", key="scd_host")
            port = st.text_input("Port", value=(
                "3306" if scd_kind == "MySQL" else "5432" if scd_kind == "PostgreSQL" else "1433" if scd_kind == "SQL Server" else "1521"),
                                key="scd_port")
            user = st.text_input("User", key="scd_user")
            pwd = st.text_input("Password", type="password", key="scd_pass")
            sid_service = st.text_input("SID/Service (Oracle)", key="scd_sid")
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""

        database = st.text_input("Database / Catalog", key="scd_db")
        table_old = st.text_input("Old snapshot table or SELECT SQL", key="scd_table_old")
        table_new = st.text_input("New snapshot table or SELECT SQL", key="scd_table_new")
        scd_type = st.selectbox("SCD Type", ["Type 1", "Type 2"], key="scd_type_select")

        if st.button("Load Snapshots and Run Validation"):
            try:
                url = build_conn_url(
                    scd_kind, host, port, user, pwd, database, sid_service,
                    aws_access_key_id=aws_access_key_id,
                    aws_secret_access_key=aws_secret_access_key,
                    aws_session_token=aws_session_token,
                    region=region,
                    s3_staging_dir=s3_staging_dir,
                    http_path=databricks_http_path
                )
                eng, err = try_connect(url)
                if not eng:
                    st.error(f"DB connection failed: {err}")
                    st.stop()

                q_old = table_old if table_old.strip().lower().startswith("select") else f"SELECT * FROM {table_old}"
                q_new = table_new if table_new.strip().lower().startswith("select") else f"SELECT * FROM {table_new}"

                df_old = pd.read_sql(text(q_old), eng)
                df_new = pd.read_sql(text(q_new), eng)

                st.session_state["scd_df_old"] = df_old
                st.session_state["scd_df_new"] = df_new

                st.success("Snapshots loaded successfully.")
            except Exception as e:
                st.error(f"Failed to load snapshots: {e}")

    else:
        # File upload mode
        col1, col2 = st.columns(2)
        with col1:
            old_file = st.file_uploader("Old snapshot file", type=["csv", "parquet", "xlsx", "json", "txt"], key="scd_old_file")
        with col2:
            new_file = st.file_uploader("New snapshot file", type=["csv", "parquet", "xlsx", "json", "txt"], key="scd_new_file")

        scd_type = st.selectbox("SCD Type", ["Type 1", "Type 2"], key="scd_type_file")

        if st.button("Load Files and Run Validation"):
            if old_file is None or new_file is None:
                st.error("Please upload both old and new snapshot files.")
                st.stop()
            try:
                df_old = read_any_file(old_file)
                df_new = read_any_file(new_file)
                st.session_state["scd_df_old"] = df_old
                st.session_state["scd_df_new"] = df_new
                st.success("Files loaded successfully.")
            except Exception as e:
                st.error(f"Failed to read files: {e}")

    # Use loaded dataframes from session state
    df_old = st.session_state.get("scd_df_old")
    df_new = st.session_state.get("scd_df_new")

    if df_old is not None:
        st.subheader("Old Snapshot Preview")
        st.dataframe(df_old.head(100))
    else:
        st.info("Old snapshot not loaded yet.")

    if df_new is not None:
        st.subheader("New Snapshot Preview")
        st.dataframe(df_new.head(100))

        if scd_type == "Type 1":
            # Type 1 validation UI and results
            null_counts = df_new.isnull().sum().sort_values(ascending=False)
            st.subheader("Null counts in new snapshot")
            st.dataframe(null_counts.reset_index().rename(columns={"index": "column", 0: "null_count"}))
            dup_rows = df_new.duplicated().sum()
            st.metric("Duplicate rows (full row) in new snapshot", dup_rows)

        else:
            # Type 2 validation UI with dynamic column selection
            cols = list(df_new.columns)
            start_col = st.selectbox("Start date column", options=cols, key="scd_start_col")
            end_col = st.selectbox("End date column", options=cols, key="scd_end_col")
            cur_flag = st.selectbox("Current flag column", options=cols, key="scd_cur_flag")

            # User must select business key columns explicitly (no default)
            bk_cols = st.multiselect("Business key columns", options=cols)

            if not bk_cols:
                st.warning("Please select at least one business key column.")
                st.stop()

            if st.button("Run SCD Type 2 Validation"):
                dfx = df_new.copy()
                for c in [start_col, end_col]:
                    dfx[c] = pd.to_datetime(dfx[c], errors="coerce")

                if cur_flag not in dfx.columns:
                    st.error(f"Current flag column '{cur_flag}' not found in new snapshot")
                    st.stop()

                # Duplicate current flags check
                dup_current = dfx[dfx[cur_flag] == True].groupby(bk_cols).size().reset_index(name="current_count")
                dup_current = dup_current[dup_current["current_count"] > 1]
                st.subheader("Duplicate current flags by business key")
                st.dataframe(dup_current)
                st.session_state["scd_dup_current"] = dup_current

                # Overlapping periods check
                overlaps = []
                for bk, grp in dfx.groupby(bk_cols):
                    grp = grp.sort_values(start_col)
                    prev_end = None
                    for _, r in grp.iterrows():
                        st_dt = r.get(start_col)
                        en_dt = r.get(end_col)
                        if prev_end is not None and st_dt is not None and en_dt is not None and st_dt < prev_end:
                            overlaps.append({"business_key": str(bk), "start": st_dt, "end": en_dt, "prev_end": prev_end})
                        if en_dt is not None:
                            prev_end = en_dt
                overlaps_df = pd.DataFrame(overlaps)
                st.subheader("Overlapping periods")
                st.dataframe(overlaps_df)
                st.session_state["scd_overlaps"] = overlaps_df

        # Display stored results persistently
        if "scd_dup_current" in st.session_state and st.session_state["scd_dup_current"] is not None and not st.session_state["scd_dup_current"].empty:
            st.subheader("Duplicate current flags by business key (previous run)")
            st.dataframe(st.session_state["scd_dup_current"])

        if "scd_overlaps" in st.session_state and st.session_state["scd_overlaps"] is not None and not st.session_state["scd_overlaps"].empty:
            st.subheader("Overlapping periods (previous run)")
            st.dataframe(st.session_state["scd_overlaps"])
    else:
        st.info("New snapshot not loaded yet.")

# ---------------------------
# Section: Mapping Generator
# ---------------------------
elif section == "Mapping Generator":
    st.header("🔗 Mapping Generator (Source -> Target)")
    st.info(
        "Connect to source & target, then generate a mapping file. Matches are green, mismatches amber in the UI and Excel.")

    # --- Source Connection Block ---
    st.subheader("Source Database Connection")
    s_c1, s_c2, s_c3 = st.columns(3)
    s_db_type = s_c1.selectbox(
        "Source DB Type",
        ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
        key="map_src_db_type"
    )

    # Initialize all source variables
    s_host = s_port = s_user = s_pass = s_db = s_sid_service = s_http_path = ""
    s_aws_access_key_id = s_aws_secret_access_key = s_aws_session_token = ""
    s3_staging_dir = ""
    s_region = ""

    if s_db_type == "Athena":
        st.markdown("### Athena Source Connection Details")
        s_aws_access_key_id = s_c1.text_input("AWS Access Key ID (Source)", type="password", key="map_src_athena_key")
        s_aws_secret_access_key = s_c2.text_input("AWS Secret Access Key (Source)", type="password", key="map_src_athena_secret")
        s_aws_session_token = s_c3.text_input("AWS Session Token (Source, optional)", type="password", key="map_src_athena_token")
        s_region = s_c1.text_input("AWS Region (Source)", value="us-east-1", key="map_src_athena_region")
        s3_staging_dir = s_c2.text_input("S3 Staging Dir (Source)", key="map_src_athena_s3")
        s_db = s_c3.text_input("Source Database / Catalog", key="map_src_db")

        s_host = f"athena.{s_region}.amazonaws.com"
        s_port = "443"
        s_user = s_aws_access_key_id
        s_pass = s_aws_secret_access_key
        s_sid_service = s3_staging_dir
        s_http_path = ""

    elif s_db_type == "Databricks":
        st.markdown("### Databricks Source Connection Details")
        s_host = s_c1.text_input("Workspace Hostname (Source)", key="map_src_databricks_host")
        s_http_path = s_c2.text_input("HTTP Path (Source)", key="map_src_databricks_http_path")
        s_pass = s_c3.text_input("Personal Access Token (Source)", type="password", key="map_src_databricks_token")
        s_user = "token"
        s_port = "443"
        s_sid_service = ""
        s_aws_access_key_id = ""
        s_aws_secret_access_key = ""
        s_aws_session_token = ""
        s3_staging_dir = ""
        s_db = s_c3.text_input("Source Database / Catalog", key="map_src_db")

    else:
        s_host = s_c1.text_input("Source Host", key="map_src_host")
        s_port = s_c2.text_input("Source Port", value="3306" if s_db_type == "MySQL" else "5432", key="map_src_port")
        s_user = s_c1.text_input("Source User", key="map_src_user")
        s_pass = s_c2.text_input("Source Password", type="password", key="map_src_pass")
        s_db = s_c3.text_input("Source Database", key="map_src_db")
        s_sid_service = s_c3.text_input("Source SID/Service (Oracle)", key="map_src_sid")
        s_http_path = ""

    if s_c3.button("Connect Source", key="map_src_connect"):
        try:
            src_url = build_conn_url(
                s_db_type, s_host, s_port, s_user, s_pass, s_db, s_sid_service,
                aws_access_key_id=s_aws_access_key_id,
                aws_secret_access_key=s_aws_secret_access_key,
                aws_session_token=s_aws_session_token,
                region=s_region if s_db_type == "Athena" else "",
                s3_staging_dir=s3_staging_dir if s_db_type == "Athena" else "",
                http_path=s_http_path if s_db_type.lower() == "databricks" else ""
            )
            src_engine, err = try_connect(src_url)
            if src_engine:
                st.session_state["map_src_engine"] = src_engine
                st.success("Source connected")
            else:
                st.error(f"Source connection failed: {err}")
        except Exception as e:
            st.error(f"Source connection error: {e}")

    # --- Target Connection Block ---
    st.subheader("Target Database Connection")
    t_c1, t_c2, t_c3 = st.columns(3)
    t_db_type = t_c1.selectbox(
        "Target DB Type",
        ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
        key="map_tgt_db_type"
    )

    # Initialize all target variables
    t_host = t_port = t_user = t_pass = t_db = t_sid_service = t_http_path = ""
    t_aws_access_key_id = t_aws_secret_access_key = t_aws_session_token = ""
    t3_staging_dir = ""
    t_region = ""

    if t_db_type == "Athena":
        st.markdown("### Athena Target Connection Details")
        t_aws_access_key_id = t_c1.text_input("AWS Access Key ID (Target)", type="password", key="map_tgt_athena_key")
        t_aws_secret_access_key = t_c2.text_input("AWS Secret Access Key (Target)", type="password", key="map_tgt_athena_secret")
        t_aws_session_token = t_c3.text_input("AWS Session Token (Target, optional)", type="password", key="map_tgt_athena_token")
        t_region = t_c1.text_input("AWS Region (Target)", value="us-east-1", key="map_tgt_athena_region")
        t3_staging_dir = t_c2.text_input("S3 Staging Dir (Target)", key="map_tgt_athena_s3")
        t_db = t_c3.text_input("Target Database / Catalog", key="map_tgt_db")

        t_host = f"athena.{t_region}.amazonaws.com"
        t_port = "443"
        t_user = t_aws_access_key_id
        t_pass = t_aws_secret_access_key
        t_sid_service = t3_staging_dir
        t_http_path = ""

    elif t_db_type == "Databricks":
        st.markdown("### Databricks Target Connection Details")
        t_host = t_c1.text_input("Workspace Hostname (Target)", key="map_tgt_databricks_host")
        t_http_path = t_c2.text_input("HTTP Path (Target)", key="map_tgt_databricks_http_path")
        t_pass = t_c3.text_input("Personal Access Token (Target)", type="password", key="map_tgt_databricks_token")
        t_user = "token"
        t_port = "443"
        t_sid_service = ""
        t_aws_access_key_id = ""
        t_aws_secret_access_key = ""
        t_aws_session_token = ""
        t3_staging_dir = ""
        t_db = t_c3.text_input("Target Database / Catalog", key="map_tgt_db")

    else:
        t_host = t_c1.text_input("Target Host", key="map_tgt_host")
        t_port = t_c2.text_input("Target Port", value="3306" if t_db_type == "MySQL" else "5432", key="map_tgt_port")
        t_user = t_c1.text_input("Target User", key="map_tgt_user")
        t_pass = t_c2.text_input("Target Password", type="password", key="map_tgt_pass")
        t_db = t_c3.text_input("Target Database", key="map_tgt_db")
        t_sid_service = t_c3.text_input("Target SID/Service (Oracle)", key="map_tgt_sid")
        t_http_path = ""

    if t_c3.button("Connect Target", key="map_tgt_connect"):
        try:
            tgt_url = build_conn_url(
                t_db_type, t_host, t_port, t_user, t_pass, t_db, t_sid_service,
                aws_access_key_id=t_aws_access_key_id,
                aws_secret_access_key=t_aws_secret_access_key,
                aws_session_token=t_aws_session_token,
                region=t_region if t_db_type == "Athena" else "",
                s3_staging_dir=t3_staging_dir if t_db_type == "Athena" else "",
                http_path=t_http_path if t_db_type.lower() == "databricks" else ""
            )
            tgt_engine, err = try_connect(tgt_url)
            if tgt_engine:
                st.session_state["map_tgt_engine"] = tgt_engine
                st.success("Target connected")
            else:
                st.error(f"Target connection failed: {err}")
        except Exception as e:
            st.error(f"Target connection error: {e}")

    # --- After connections, input table names and generate mapping ---
    src_engine = st.session_state.get("map_src_engine")
    tgt_engine = st.session_state.get("map_tgt_engine")

    if src_engine and tgt_engine:
        st.markdown("---")
        s_table = st.text_input("Source Table Name", key="map_src_table_name")
        t_table = st.text_input("Target Table Name", key="map_tgt_table_name")

        if st.button("Generate Mapping"):
            if not s_table or not t_table:
                st.error("Please provide both source and target table names.")
            else:
                try:
                    src_insp = inspect(src_engine)
                    tgt_insp = inspect(tgt_engine)

                    src_cols = src_insp.get_columns(s_table)
                    tgt_cols = tgt_insp.get_columns(t_table)

                    df_src = pd.DataFrame({
                        "column_name": [c["name"] for c in src_cols],
                        "data_type_src": [str(c["type"]) for c in src_cols]
                    })
                    df_tgt = pd.DataFrame({
                        "column_name": [c["name"] for c in tgt_cols],
                        "data_type_tgt": [str(c["type"]) for c in tgt_cols]
                    })

                    df_map = pd.merge(df_src, df_tgt, on="column_name", how="outer")

                    df_map["match"] = df_map.apply(
                        lambda r: str(r.get("data_type_src", "")).lower() == str(r.get("data_type_tgt", "")).lower()
                        if pd.notna(r.get("data_type_src")) and pd.notna(r.get("data_type_tgt")) else False,
                        axis=1
                    )

                    def color_row(row):
                        color = "background-color: #c6efce" if row.match else "background-color: #ffeb9c"
                        return [color] * len(row)

                    st.markdown("### Mapping Results (green = match, amber = mismatch)")
                    st.dataframe(df_map.style.apply(color_row, axis=1), use_container_width=True)

                    excel_bytes = export_mapping_to_excel(df_map)
                    st.download_button("Download Mapping Excel", data=excel_bytes, file_name=f"mapping_{int(time.time())}.xlsx")

                except Exception as e:
                    st.error(f"Mapping generation failed: {e}")
    else:
        st.info("Connect both Source and Target databases to enable mapping generation.")
# ---------------------------
# Section: Data Quality
# ---------------------------
elif section == "Data Quality":
    st.header("🧪 Data Quality Checks")
    st.markdown("Select rules to validate data quality. Compact UI with all checks and detailed results.")

    # Load source
    src_choice = st.radio("Source type", ["File", "Database"], horizontal=True, key="dq_source")
    dq_df = None

    if src_choice == "File":
        dq_file = st.file_uploader("Upload file", type=["csv", "xlsx", "json", "parquet", "txt"], key="dq_upload")
        if dq_file:
            dq_df = read_any_file(dq_file)
    else:
        c1, c2 = st.columns(2)
        dq_dbtype = c1.selectbox("DB Type", ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"], key="dq_dbtype")
        if dq_dbtype == "Athena":
            st.markdown("### Athena Connection Details")
            aws_access_key_id = c1.text_input("AWS Access Key ID", type="password", key="dq_athena_key")
            aws_secret_access_key = c2.text_input("AWS Secret Access Key", type="password", key="dq_athena_secret")
            aws_session_token = c2.text_input("AWS Session Token (optional)", type="password", key="dq_athena_token")
            region = c1.text_input("AWS Region", value="us-east-1", key="dq_athena_region")
            s3_staging_dir = c2.text_input("S3 Staging Dir (s3://bucket/folder/)", key="dq_athena_s3")
            dq_host = f"athena.{region}.amazonaws.com"
            dq_port = "443"
            dq_user = aws_access_key_id
            dq_pwd = aws_secret_access_key
            dq_sid_service = s3_staging_dir
        elif dq_dbtype == "Databricks":
            st.markdown("### Databricks Connection Details")
            databricks_host = st.text_input("Workspace Hostname", key="single_db_databricks_host")
            databricks_http_path = st.text_input("HTTP Path", key="single_db_databricks_http_path")
            databricks_token = st.text_input("Personal Access Token", type="password", key="single_db_databricks_token")
            database = st.text_input("Database / Catalog", key="single_db")
            dq_host = databricks_host
            dq_port = "443"
            dq_user = "token"
            dq_pwd = databricks_token
            dq_sid_service = ""
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""
            http_path = databricks_http_path
        else:
            dq_host = c1.text_input("Host", key="dq_host")
            dq_port = c2.text_input("Port", value=("3306" if dq_dbtype=="MySQL" else "5432" if dq_dbtype=="Postgres" else "1433" if dq_dbtype=="SQL Server" else "1521"), key="dq_port")
            dq_user = c1.text_input("User", key="dq_user")
            dq_pwd = c2.text_input("Password", type="password", key="dq_pwd")
            dq_sid_service = c2.text_input("SID/Service (Oracle)", key="dq_sid")
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""

        dq_db = c2.text_input("Database", key="dq_db")
        dq_table = st.text_input("Table name or SELECT SQL", key="dq_table")

        if st.button("Load DB Table", key="dq_load_db"):
            try:
                url = build_conn_url(
                    dq_dbtype, dq_host, dq_port, dq_user, dq_pwd, dq_db, dq_sid_service,
                    aws_access_key_id=aws_access_key_id,
                    aws_secret_access_key=aws_secret_access_key,
                    aws_session_token=aws_session_token,
                    region=region,
                    s3_staging_dir=s3_staging_dir,
                    http_path = http_path
                )
                eng, err = try_connect(url)
                if eng:
                    dq_df = pd.read_sql(
                        text(dq_table) if dq_table.strip().lower().startswith("select") else text(f"SELECT * FROM {dq_table}"),
                        eng
                    )
                    st.session_state["dq_df"] = dq_df
                    st.success(f"Loaded {len(dq_df)} rows, {len(dq_df.columns)} cols")
                else:
                    st.error(f"DB connection failed: {err}")
            except Exception as e:
                st.error(f"DB Load failed: {e}")
        else:
            dq_df = st.session_state.get("dq_df")

    if dq_df is not None:
        st.dataframe(dq_df.head(20), use_container_width=True)
        cols = list(dq_df.columns)

        st.markdown("### ✅ Select Data Quality Checks")
        selected_checks = st.multiselect(
            "Choose checks to run",
            ["Nulls", "Duplicates", "Distinct Count", "Range", "Length", "Pattern",
             "Equality", "Greater/Less", "Between", "Custom Expression"],
            default=["Nulls", "Duplicates", "Distinct Count"]
        )

        dq_rules = {}

        with st.expander("⚙️ Configure Checks", expanded=True):
            if "Nulls" in selected_checks:
                dq_rules["null_col"] = st.selectbox("Column for Null check", ["--none--"] + cols, key="dq_null")

            if "Duplicates" in selected_checks:
                dq_rules["dup_col"] = st.selectbox("Column for Duplicate check", ["--none--"] + cols, key="dq_dup")

            if "Distinct Count" in selected_checks:
                dq_rules["distinct_col"] = st.selectbox("Column for Distinct count", ["--none--"] + cols, key="dq_dist")

            if "Range" in selected_checks:
                c1, c2, c3 = st.columns(3)
                dq_rules["range_col"] = c1.selectbox("Range column", ["--none--"] + cols, key="dq_range")
                dq_rules["min_val"] = c2.text_input("Min", key="dq_min")
                dq_rules["max_val"] = c3.text_input("Max", key="dq_max")

            if "Length" in selected_checks:
                c1, c2, c3 = st.columns(3)
                dq_rules["len_col"] = c1.selectbox("Length column", ["--none--"] + cols, key="dq_len")
                dq_rules["min_len"] = c2.number_input("Min length", min_value=0, value=0, key="dq_min_len")
                dq_rules["max_len"] = c3.number_input("Max length", min_value=0, value=100, key="dq_max_len")

            if "Pattern" in selected_checks:
                dq_rules["pat_col"] = st.selectbox("Pattern column", ["--none--"] + cols, key="dq_pat")
                dq_rules["pattern"] = st.text_input("Regex Pattern", key="dq_regex")

            if "Equality" in selected_checks:
                c1, c2 = st.columns(2)
                dq_rules["eq_col1"] = c1.selectbox("Col1", ["--none--"] + cols, key="dq_eq1")
                dq_rules["eq_col2"] = c2.selectbox("Col2", ["--none--"] + cols, key="dq_eq2")

            if "Greater/Less" in selected_checks:
                c1, c2 = st.columns(2)
                dq_rules["gt_col1"] = c1.selectbox("Col1", ["--none--"] + cols, key="dq_gt1")
                dq_rules["gt_col2"] = c2.selectbox("Col2", ["--none--"] + cols, key="dq_gt2")

            if "Between" in selected_checks:
                c1, c2, c3 = st.columns(3)
                dq_rules["between_col"] = c1.selectbox("Column", ["--none--"] + cols, key="dq_between")
                dq_rules["between_min"] = c2.text_input("Min", key="dq_between_min")
                dq_rules["between_max"] = c3.text_input("Max", key="dq_between_max")

            if "Custom Expression" in selected_checks:
                dq_rules["custom_expr"] = st.text_area("Custom Expression (use df)", key="dq_custom")

        if st.button("🚀 Run Checks"):
            summary, violations = [], {}

            # --- Null check
            if dq_rules.get("null_col") and dq_rules["null_col"] != "--none--":
                failed = dq_df[dq_df[dq_rules["null_col"]].isna()]
                summary.append({"Rule": f"Nulls in {dq_rules['null_col']}", "Failed": len(failed)})
                if not failed.empty: violations["Nulls"] = failed

            # --- Duplicates
            if dq_rules.get("dup_col") and dq_rules["dup_col"] != "--none--":
                dup_mask = dq_df.duplicated(subset=[dq_rules["dup_col"]], keep=False)
                failed = dq_df[dup_mask]
                summary.append({"Rule": f"Duplicates in {dq_rules['dup_col']}", "Failed": len(failed)})
                if not failed.empty: violations["Duplicates"] = failed

            # --- Distinct
            if dq_rules.get("distinct_col") and dq_rules["distinct_col"] != "--none--":
                count = dq_df[dq_rules["distinct_col"]].nunique(dropna=True)
                summary.append({"Rule": f"Distinct values in {dq_rules['distinct_col']}", "Failed": f"{count} unique"})

            # --- Range
            if dq_rules.get("range_col") and dq_rules["range_col"] != "--none--":
                vals = pd.to_numeric(dq_df[dq_rules["range_col"]], errors="coerce")
                mask = (vals < float(dq_rules["min_val"])) | (vals > float(dq_rules["max_val"]))
                failed = dq_df[mask]
                summary.append(
                    {"Rule": f"Range [{dq_rules['min_val']}, {dq_rules['max_val']}] on {dq_rules['range_col']}",
                     "Failed": len(failed)})
                if not failed.empty: violations["Range"] = failed

            # --- Length
            if dq_rules.get("len_col") and dq_rules["len_col"] != "--none--":
                lens = dq_df[dq_rules["len_col"]].astype(str).str.len()
                mask = (lens < dq_rules["min_len"]) | (lens > dq_rules["max_len"])
                failed = dq_df[mask]
                summary.append({"Rule": f"Length {dq_rules['min_len']}–{dq_rules['max_len']} on {dq_rules['len_col']}",
                                "Failed": len(failed)})
                if not failed.empty: violations["Length"] = failed

            # --- Pattern
            if dq_rules.get("pat_col") and dq_rules["pat_col"] != "--none--" and dq_rules.get("pattern"):
                mask = dq_df[dq_rules["pat_col"]].astype(str).str.match(dq_rules["pattern"], na=False)
                failed = dq_df[~mask]
                summary.append(
                    {"Rule": f"Pattern {dq_rules['pattern']} on {dq_rules['pat_col']}", "Failed": len(failed)})
                if not failed.empty: violations["Pattern"] = failed

            # --- Equality
            eq_col1 = dq_rules.get("eq_col1")
            eq_col2 = dq_rules.get("eq_col2")
            if eq_col1 and eq_col2 and eq_col1 != "--none--" and eq_col2 != "--none--":
                mask = dq_df[eq_col1] != dq_df[eq_col2]
                failed = dq_df[mask]
                summary.append({"Rule": f"{eq_col1} == {eq_col2}", "Failed": len(failed)})
                if not failed.empty: violations["Equality"] = failed

            # --- Greater/Less
            gt_col1 = dq_rules.get("gt_col1")
            gt_col2 = dq_rules.get("gt_col2")
            if gt_col1 and gt_col2 and gt_col1 != "--none--" and gt_col2 != "--none--":
                mask = dq_df[gt_col1] <= dq_df[gt_col2]
                failed = dq_df[mask]
                summary.append({"Rule": f"{gt_col1} > {gt_col2}", "Failed": len(failed)})
                if not failed.empty: violations["Greater/Less"] = failed

            # --- Between
            if dq_rules.get("between_col") and dq_rules["between_col"] != "--none--":
                vals = pd.to_numeric(dq_df[dq_rules["between_col"]], errors="coerce")
                mask = (vals < float(dq_rules["between_min"])) | (vals > float(dq_rules["between_max"]))
                failed = dq_df[mask]
                summary.append(
                    {"Rule": f"{dq_rules['between_min']} <= {dq_rules['between_col']} <= {dq_rules['between_max']}",
                     "Failed": len(failed)})
                if not failed.empty: violations["Between"] = failed

            # --- Custom Expression
            if dq_rules.get("custom_expr"):
                try:
                    result = eval(dq_rules["custom_expr"], {"df": dq_df, "pd": pd, "np": np})
                    if isinstance(result, pd.Series):
                        failed = dq_df[~result]
                        summary.append({"Rule": f"Custom {dq_rules['custom_expr']}", "Failed": len(failed)})
                        if not failed.empty: violations["Custom"] = failed
                except Exception as e:
                    summary.append({"Rule": "Custom Expression ERROR", "Failed": str(e)})

            # Show summary
            st.subheader("📊 DQ Summary")
            df_summary = pd.DataFrame(summary)
            st.dataframe(df_summary, use_container_width=True)

            # Show failed records
            for rule, failed_df in violations.items():
                with st.expander(f"❌ {rule} Violations ({len(failed_df)})"):
                    st.dataframe(failed_df.head(100), use_container_width=True)
                    st.download_button(f"Download {rule} Violations", failed_df.to_csv(index=False).encode("utf-8"),
                                       file_name=f"dq_{rule}_violations.csv")

            st.download_button("📥 Download Summary", df_summary.to_csv(index=False).encode("utf-8"),
                               file_name="dq_summary.csv")
            _log_entry("Data Quality", "Checks run", result_summary={"summary": summary})
# ---------------------------
# Section: Reconciliation
# ---------------------------
elif section == "Reconciliation":
    st.header("Reconciliation")
    st.markdown(
        "Aggregate reconciliation between two independent sources (DB↔DB, DB↔File, or File↔File). "
        "Connect source and target independently and run aggregate operations like count, sum, min, max, null_count."
    )

    # Helper: DB connection form
    def db_connection_form(prefix: str):
        st.subheader(f"{prefix} Database Connection")
        c1, c2, c3 = st.columns(3)
        db_type = c1.selectbox(f"{prefix} DB Type",
                              ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
                              key=f"recon_{prefix.lower()}_db_type")
        if db_type == "Athena":
            st.markdown(f"### Athena {prefix} Details")
            aws_access_key_id = c1.text_input(f"AWS Access Key ID ({prefix})", type="password", key=f"recon_{prefix.lower()}_athena_key")
            aws_secret_access_key = c2.text_input(f"AWS Secret Access Key ({prefix})", type="password", key=f"recon_{prefix.lower()}_athena_secret")
            aws_session_token = c3.text_input(f"AWS Session Token ({prefix}, optional)", type="password", key=f"recon_{prefix.lower()}_athena_token")
            region = c1.text_input(f"AWS Region ({prefix})", value="us-east-1", key=f"recon_{prefix.lower()}_athena_region")
            s3_staging_dir = c2.text_input(f"S3 Staging Dir ({prefix})", key=f"recon_{prefix.lower()}_athena_s3")
            host = f"athena.{region}.amazonaws.com"
            port = "443"
            user = aws_access_key_id
            pwd = aws_secret_access_key
            database = st.text_input(f"{prefix} Database / Catalog", key=f"recon_{prefix.lower()}_db")
            sid_service = s3_staging_dir
        elif db_type == "Databricks":
            st.markdown("### Databricks Connection Details")
            databricks_host = st.text_input("Workspace Hostname", key=f"recon_{prefix.lower()}_databricks_host")
            databricks_http_path = st.text_input("HTTP Path", key=f"recon_{prefix.lower()}_databricks_http_path")
            databricks_token = st.text_input("Personal Access Token", type="password", key=f"recon_{prefix.lower()}_databricks_token")
            database = st.text_input("Database / Catalog", key=f"recon_{prefix.lower()}_database")
            host = databricks_host
            port = "443"
            user = "token"
            pwd = databricks_token
            sid_service = ""
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""
            http_path = databricks_http_path
        else:
            host = c1.text_input(f"{prefix} Host", key=f"recon_{prefix.lower()}_host")
            port = c2.text_input(f"{prefix} Port", value=(
                "3306" if db_type == "MySQL" else "5432" if db_type == "Postgres" else "1433" if db_type == "SQL Server" else "1521"),
                                 key=f"recon_{prefix.lower()}_port")
            user = c1.text_input(f"{prefix} User", key=f"recon_{prefix.lower()}_user")
            pwd = c2.text_input(f"{prefix} Password", type="password", key=f"recon_{prefix.lower()}_pwd")
            database = c3.text_input(f"{prefix} Database / Catalog", key=f"recon_{prefix.lower()}_db")
            sid_service = c3.text_input(f"{prefix} SID/Service (Oracle)", key=f"recon_{prefix.lower()}_sid")
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""

        if st.button(f"Connect {prefix} DB", key=f"recon_{prefix.lower()}_connect"):
            try:
                url = build_conn_url(
                    db_type, host, port, user, pwd, database, sid_service,
                    aws_access_key_id=aws_access_key_id,
                    aws_secret_access_key=aws_secret_access_key,
                    aws_session_token=aws_session_token,
                    region=region,
                    s3_staging_dir=s3_staging_dir,
                    http_path=databricks_http_path
                )
                eng, err = try_connect(url)
                if eng:
                    st.session_state[f"recon_{prefix.lower()}_engine"] = eng
                    st.success(f"{prefix} DB connected")
                else:
                    st.error(f"{prefix} DB connection failed: {err}")
            except Exception as e:
                st.error(f"{prefix} DB connection error: {e}")

    # Show connection forms
    db_connection_form("Source")
    db_connection_form("Target")

    # Load data from source and target
    def load_data(prefix: str):
        st.subheader(f"{prefix} Data Load")
        choice = st.selectbox(f"{prefix} source type", ["File", "DB"], key=f"recon_{prefix.lower()}_source_type")
        df = None
        if choice == "File":
            f = st.file_uploader(f"Upload {prefix} file", type=["csv", "json", "parquet", "xlsx", "txt"], key=f"recon_{prefix.lower()}_file")
            if f:
                df = read_any_file(f)
                st.session_state[f"recon_{prefix.lower()}_df"] = df
                st.success(f"{prefix} file loaded")
                st.dataframe(df.head(20))
        else:
            tbl = st.text_input(f"{prefix} DB table or SQL SELECT", key=f"recon_{prefix.lower()}_table")
            if st.button(f"Load {prefix} from DB", key=f"recon_{prefix.lower()}_load_db"):
                eng = st.session_state.get(f"recon_{prefix.lower()}_engine")
                if not eng:
                    st.error(f"No {prefix} DB connection found. Connect first.")
                elif not tbl:
                    st.error(f"Provide {prefix} table name or SQL SELECT.")
                else:
                    try:
                        if tbl.strip().lower().startswith("select"):
                            df = pd.read_sql(text(tbl), eng)
                        else:
                            df = pd.read_sql(text(f"SELECT * FROM {tbl} LIMIT 10000"), eng)
                        st.session_state[f"recon_{prefix.lower()}_df"] = df
                        st.success(f"{prefix} data loaded ({len(df)} rows)")
                        st.dataframe(df.head(20))
                    except Exception as e:
                        st.error(f"Failed to load {prefix} data: {e}")
        return st.session_state.get(f"recon_{prefix.lower()}_df")

    left_df = load_data("Source")
    right_df = load_data("Target")

    # Aggregate reconciliation UI
    st.markdown("---")
    st.subheader("Aggregate Reconciliation Operations")
    st.markdown(
        "Enter one aggregate operation per line. Supported operations:\n"
        "- count(*)\n"
        "- sum(column_name)\n"
        "- min(column_name)\n"
        "- max(column_name)\n"
        "- null_count(column_name)\n"
        "Example:\n"
        "count(*)\nsum(amount)\nmin(price)\nmax(price)\nnull_count(id)"
    )
    recon_ops = st.text_area("Operations", value="count(*)", key="recon_ops")

    if st.button("Run Aggregate Reconciliation", key="recon_agg_run"):
        if left_df is None or right_df is None:
            st.error("Load both Source and Target data first.")
        else:
            try:
                ops = [x.strip() for x in recon_ops.splitlines() if x.strip()]

                def run_agg_ops(df, ops):
                    out = {}
                    for op in ops:
                        op_lower = op.lower()
                        if op_lower == "count(*)":
                            out[op] = len(df)
                        elif op_lower.startswith("sum(") and op.endswith(")"):
                            col = op[4:-1].strip()
                            if col in df.columns:
                                out[op] = float(pd.to_numeric(df[col], errors="coerce").sum())
                            else:
                                out[op] = "col_missing"
                        elif op_lower.startswith("min(") and op.endswith(")"):
                            col = op[4:-1].strip()
                            if col in df.columns:
                                out[op] = df[col].min()
                            else:
                                out[op] = "col_missing"
                        elif op_lower.startswith("max(") and op.endswith(")"):
                            col = op[4:-1].strip()
                            if col in df.columns:
                                out[op] = df[col].max()
                            else:
                                out[op] = "col_missing"
                        elif op_lower.startswith("null_count(") and op.endswith(")"):
                            col = op[11:-1].strip()
                            if col in df.columns:
                                out[op] = int(df[col].isna().sum())
                            else:
                                out[op] = "col_missing"
                        else:
                            out[op] = "unsupported"
                    return out

                left_res = run_agg_ops(left_df, ops)
                right_res = run_agg_ops(right_df, ops)
                rows = []
                for op in ops:
                    rows.append({
                        "operation": op,
                        "source": left_res.get(op),
                        "target": right_res.get(op),
                        "match": left_res.get(op) == right_res.get(op)
                    })
                df_agg = pd.DataFrame(rows)

                st.success("Aggregate reconciliation completed.")
                st.dataframe(df_agg, use_container_width=True)

                st.download_button(
                    "Download Aggregate Reconciliation CSV",
                    df_agg.to_csv(index=False).encode("utf-8"),
                    file_name="reconciliation_aggregate.csv"
                )

                _log_entry("Reconciliation", "run aggregate reconciliation",
                           result_summary={"aggregate_ops": len(ops), "results": df_agg.to_dict(orient="records")})

            except Exception as e:
                st.error(f"Aggregate reconciliation failed: {e}")

# ---------------------------
# Section: Profiling
# ---------------------------
elif section == "Profiling":
    st.header("Profiling")
    st.markdown("Enhanced profiling with data completeness, unique value counts, outlier detection, and data type summary.")

    prof_choice = st.selectbox("Load data for profiling from", ["File", "DB"], key="prof_choice")
    prof_df = None

    if prof_choice == "File":
        pf = st.file_uploader("Upload file", type=["csv", "parquet", "xlsx", "json", "txt"], key="prof_file")
        if pf:
            prof_df = read_any_file(pf)
            st.session_state["prof_df"] = prof_df
    else:
        # Independent DB connection form for profiling (like other sections)
        c1, c2, c3 = st.columns(3)
        db_type = c1.selectbox("DB Type",
                              ["MySQL", "Postgres", "SQL Server", "Oracle", "Redshift", "Athena", "Databricks"],
                              key="prof_db_type")
        if db_type == "Athena":
            st.markdown("### Athena Connection Details")
            aws_access_key_id = c1.text_input("AWS Access Key ID", type="password", key="prof_athena_key")
            aws_secret_access_key = c2.text_input("AWS Secret Access Key", type="password", key="prof_athena_secret")
            aws_session_token = c3.text_input("AWS Session Token (optional)", type="password", key="prof_athena_token")
            region = c1.text_input("AWS Region", value="us-east-1", key="prof_athena_region")
            s3_staging_dir = c2.text_input("S3 Staging Dir (s3://bucket/folder/)", key="prof_athena_s3")
            host = f"athena.{region}.amazonaws.com"
            port = "443"
            user = aws_access_key_id
            pwd = aws_secret_access_key
            database = st.text_input("Database / Catalog", key="prof_db")
            sid_service = s3_staging_dir
        elif db_type == "Databricks":
            st.markdown("### Databricks Connection Details")
            databricks_host = st.text_input("Workspace Hostname", key="prof_databricks_host")
            databricks_http_path = st.text_input("HTTP Path", key="prof_databricks_http_path")
            databricks_token = st.text_input("Personal Access Token", type="password",
                                             key="prof_databricks_token")
            database = st.text_input("Database / Catalog", key="prof_database")
            host = databricks_host
            port = "443"
            user = "token"
            pwd = databricks_token
            sid_service = ""
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""
            http_path = databricks_http_path
        else:
            host = c1.text_input("Host", key="prof_host")
            port = c2.text_input("Port", value=(
                "3306" if db_type == "MySQL" else "5432" if db_type == "Postgres" else "1433" if db_type == "SQL Server" else "1521"),
                                 key="prof_port")
            user = c1.text_input("User", key="prof_user")
            pwd = c2.text_input("Password", type="password", key="prof_pwd")
            database = c3.text_input("Database / Catalog", key="prof_db")
            sid_service = c3.text_input("SID/Service (Oracle)", key="prof_sid")
            aws_access_key_id = ""
            aws_secret_access_key = ""
            aws_session_token = ""
            region = ""
            s3_staging_dir = ""

        # Table name input outside the button block to retain value
        tbl = st.text_input("Table name to profile", key="prof_table_name")

        if st.button("Connect and Load Table", key="prof_load"):
            if not tbl:
                st.error("Provide table name")
            else:
                try:
                    url = build_conn_url(
                        db_type, host, port, user, pwd, database, sid_service,
                        aws_access_key_id=aws_access_key_id,
                        aws_secret_access_key=aws_secret_access_key,
                        aws_session_token=aws_session_token,
                        region=region,
                        s3_staging_dir=s3_staging_dir,
                        http_path=databricks_http_path
                    )
                    eng, err = try_connect(url)
                    if eng:
                        prof_df = pd.read_sql(text(f"SELECT * FROM {tbl} LIMIT 50000"), eng)
                        st.session_state["prof_df"] = prof_df
                        st.success(f"Loaded {len(prof_df)} rows, {len(prof_df.columns)} columns")
                    else:
                        st.error(f"Connection failed: {err}")
                except Exception as e:
                    st.error(f"Load failed: {e}")

    prof_df = st.session_state.get("prof_df")
    if prof_df is not None:
        st.write(f"Rows: {len(prof_df)}, Columns: {len(prof_df.columns)}")
        st.dataframe(prof_df.head(50))

        # Data Completeness Percentage per column
        st.subheader("Data Completeness (%) per Column")
        completeness = 100 * prof_df.notnull().mean()
        completeness_df = completeness.reset_index()
        completeness_df.columns = ["Column", "Completeness (%)"]
        completeness_df["Completeness (%)"] = completeness_df["Completeness (%)"].round(2)
        st.dataframe(completeness_df, use_container_width=True)

        # Unique Value Counts with Top Frequent Values
        st.subheader("Unique Value Counts and Top 5 Frequent Values per Column")
        unique_summary = []
        for col in prof_df.columns:
            unique_count = prof_df[col].nunique(dropna=True)
            top_vals = prof_df[col].value_counts(dropna=True).head(5)
            top_vals_str = ", ".join([f"{str(idx)} ({cnt})" for idx, cnt in top_vals.items()])
            unique_summary.append({"Column": col, "Unique Count": unique_count, "Top 5 Values (count)": top_vals_str})
        unique_df = pd.DataFrame(unique_summary)
        st.dataframe(unique_df, use_container_width=True)

        # Data Type Summary
        st.subheader("Data Type Summary")
        dtype_counts = prof_df.dtypes.value_counts().reset_index()
        dtype_counts.columns = ["Data Type", "Count"]
        st.dataframe(dtype_counts, use_container_width=True)

        # Download profiling summaries as CSV
        st.download_button("Download Data Completeness CSV", completeness_df.to_csv(index=False).encode("utf-8"),
                           file_name="profiling_data_completeness.csv")
        st.download_button("Download Unique Values CSV", unique_df.to_csv(index=False).encode("utf-8"),
                           file_name="profiling_unique_values.csv")
        st.download_button("Download Data Type Summary CSV", dtype_counts.to_csv(index=False).encode("utf-8"),
                           file_name="profiling_data_types.csv")

        _log_entry("Profiling", "Enhanced profiling run", result_summary={
            "rows": len(prof_df),
            "columns": len(prof_df.columns)
        })
# -------------------------
# SECTION: DBT (Local CLI)
# -------------------------
elif section == "DBT":
    st.header("🛠️ DBT Integration (Local CLI)")
    st.info(
        "You need dbt CLI installed locally and in PATH. Profiles will be written to profiles.yml path you provide (e.g., ~/.dbt/profiles.yml).")

    st.subheader("DBT Project Setup")
    project_mode = st.selectbox("Mode", ["Setup existing project", "Create new project"])
    project_path = st.text_input("DBT project folder path (local)", value="",
                                 help="Folder where dbt project exists or will be created")
    profiles_path = st.text_input("profiles.yml path (e.g., ~/.dbt/profiles.yml)",
                                  value=os.path.expanduser("~/.dbt/profiles.yml"))
    dbt_conn_kind = st.selectbox("DB for dbt profile",
                                 ["PostgreSQL", "MySQL", "Redshift", "Snowflake", "BigQuery", "Oracle", "Athena"],
                                 index=0)
    dbt_host = st.text_input("DB host", key="dbt_host")
    dbt_port = st.text_input("DB port", value="5432", key="dbt_port")
    dbt_db = st.text_input("DB name", key="dbt_db")
    dbt_user = st.text_input("DB user", key="dbt_user")
    dbt_pwd = st.text_input("DB password", type="password", key="dbt_pwd")
    if st.button("Save dbt profiles.yml"):
        try:
            # Create a simple profile dict for core adapters (Postgres/Redshift/MySQL/Oracle)
            profile = {
                "my_profile": {
                    "outputs": {
                        "dev": {
                            "type": "postgres" if dbt_conn_kind.lower() in ("postgresql", "redshift") else (
                                "mysql" if dbt_conn_kind.lower() == "mysql" else "oracle"),
                            "host": dbt_host,
                            "user": dbt_user,
                            "pass": dbt_pwd,
                            "port": int(dbt_port) if dbt_port else None,
                            "dbname": dbt_db,
                            "schema": "public"
                        }
                    },
                    "target": "dev"
                }
            }
            os.makedirs(os.path.dirname(profiles_path), exist_ok=True)
            import yaml

            with open(profiles_path, "w") as f:
                yaml.dump(profile, f)
            st.success(f"profiles.yml written to {profiles_path}")
            log("dbt_profiles_write", {"path": profiles_path})
        except Exception as e:
            st.error(f"Failed to write profiles.yml: {e}")

    st.markdown("---")
    st.subheader("Generate DBT singular tests (SQL .sql files)")
    gen_mode = st.selectbox("Generate tests by",
                            ["Write SQL in editor (Gemini-assisted)", "Upload Excel containing SQLs"],
                            key="dbt_gen_mode")
    if gen_mode == "Write SQL in editor (Gemini-assisted)":
        prompt_dbt = st.text_area("Describe the test you want (Gemini will generate a test SQL)", height=140)
        dialect_hint = st.selectbox("Dialect hint for test SQL",
                                    ["ANSI", "PostgreSQL", "MySQL", "SQL Server", "Databricks", "Redshift", "Athena"],
                                    key="dbt_test_dialect")
        if st.button("Generate test SQL (Gemini)"):
            test_sql = generate_sql(prompt_dbt, dialect=dialect_hint)
            st.code(test_sql, language="sql")
            st.session_state["last_generated_test_sql"] = test_sql
        test_name = st.text_input("Test file name (without .sql)", value="test_singular_1")
        target_folder = st.text_input("Project folder to store tests (e.g., <project>/tests/)",
                                      value=os.path.join(project_path, "tests"))
        if st.button("Save test to file"):
            sql_content = st.session_state.get("last_generated_test_sql", "")
            if not sql_content:
                st.warning("Generate or paste SQL first.")
            else:
                os.makedirs(target_folder, exist_ok=True)
                fn = os.path.join(target_folder, f"{test_name}.sql")
                with open(fn, "w") as f:
                    f.write(sql_content)
                st.success(f"Saved test to {fn}")
                log("dbt_test_saved", {"file": fn})

    else:
        uploaded_xl = st.file_uploader("Upload Excel with test SQLs (columns: test_name, sql)", type=["xlsx", "xls"])
        if uploaded_xl:
            df = pd.read_excel(uploaded_xl)
            st.dataframe(df.head(200))
            target_folder = st.text_input("Project tests folder", value=os.path.join(project_path, "tests"),
                                          key="dbt_upload_folder")
            if st.button("Save uploaded tests"):
                os.makedirs(target_folder, exist_ok=True)
                for _, r in df.iterrows():
                    test_name = str(r.get("test_name") or f"test_{int(time.time())}")
                    sql_content = str(r.get("sql") or "")
                    if sql_content:
                        fn = os.path.join(target_folder, f"{test_name}.sql")
                        with open(fn, "w") as f:
                            f.write(sql_content)
                st.success("Uploaded tests saved.")
                log("dbt_tests_uploaded", {"count": len(df)})

    st.markdown("---")
    st.subheader("Run DBT tests & get report")
    if st.button("Run dbt test (project folder)"):
        if not project_path:
            st.warning("Provide project folder path.")
        else:
            # run: dbt test --project-dir <project_path>
            try:
                code, out, err = run_dbt_command(project_path, ["test"])
                st.text("DBT stdout:")
                st.text(out)
                if err:
                    st.text("DBT stderr:")
                    st.text(err)
                if code == 0:
                    st.success("dbt test completed successfully.")
                else:
                    st.warning(f"dbt test exited with code {code}.")
                log("dbt_test_run", {"project": project_path, "exit_code": code})
            except Exception as e:
                st.error(f"dbt run failed: {e}")

# ---------------------------
# Section: Logs & Downloads
# ---------------------------
elif section == "Logs & Downloads":
    st.header("Logs & Downloads")
    logs = st.session_state.get("logs", [])
    st.write(f"Total log entries: {len(logs)}")
    if logs:
        df_logs = pd.DataFrame(logs)
        st.dataframe(df_logs.head(200), use_container_width=True)
        st.download_button("Download logs (CSV)", df_logs.to_csv(index=False).encode("utf-8"),
                           file_name="genai_db_tool_logs.csv")
    if st.button("Clear logs"):
        st.session_state["logs"] = []
        st.success("Logs cleared")

# ---------------------------
# End of app
# ---------------------------

st.sidebar.markdown("---")
st.sidebar.caption("App generated/updated based on user's positive.txt (functions reused).")
